# /// script
# requires-python = ">=3.12"
# dependencies = ["openpyxl>=3.1,<4", "pillow>=11,<13", "click>=8.3,<9"]
# ///
"""Excel (.xlsx) operations tool.

Commands: read, write, write-range, build, info, add-sheet, export.

Usage:
    uv run xlsx_tool.py COMMAND [OPTIONS]
"""

import csv
import io
import json
import sys
from pathlib import Path

import click
from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.drawing.image import Image as SpreadsheetImage
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils.cell import coordinate_to_tuple, range_boundaries
from openpyxl.worksheet.datavalidation import DataValidation


def _json_serial(obj: object) -> str:
    """JSON serializer that uses ISO 8601 for dates/datetimes."""
    if hasattr(obj, "isoformat"):
        return obj.isoformat()
    return str(obj)


def write_output(text: str, output_path: str | None) -> None:
    """Write text to file or stdout."""
    if output_path:
        Path(output_path).write_text(text, encoding="utf-8")
        click.echo(f"Output written to {output_path}", err=True)
    else:
        click.echo(text)


def _load_json_input(value: str) -> object:
    """Load a JSON string or a JSON file path."""
    candidate = Path(value)
    if candidate.exists() and candidate.is_file():
        return json.loads(candidate.read_text(encoding="utf-8"))
    return json.loads(value)


def _color(value: object, default: str | None = None) -> str | None:
    if value is None:
        return default
    text = str(value).strip().lstrip("#")
    return text.upper() if text else default


def _apply_style(cell, style: dict[str, object]) -> None:
    """Apply the supported, serializable subset of openpyxl cell styles."""
    font = style.get("font")
    if isinstance(font, dict):
        cell.font = Font(
            name=font.get("name"),
            size=font.get("size"),
            bold=font.get("bold"),
            italic=font.get("italic"),
            underline=font.get("underline"),
            color=_color(font.get("color")),
        )
    fill = style.get("fill")
    if isinstance(fill, dict):
        cell.fill = PatternFill(
            fill_type=str(fill.get("type", "solid")),
            fgColor=_color(fill.get("color"), "FFFFFF"),
        )
    alignment = style.get("alignment")
    if isinstance(alignment, dict):
        cell.alignment = Alignment(
            horizontal=alignment.get("horizontal"),
            vertical=alignment.get("vertical"),
            wrap_text=alignment.get("wrapText"),
            text_rotation=alignment.get("textRotation", 0),
        )
    border = style.get("border")
    if isinstance(border, dict):
        def make_side(name: str) -> Side:
            value = border.get(name)
            if not isinstance(value, dict):
                return Side()
            return Side(style=value.get("style", "thin"), color=_color(value.get("color"), "000000"))
        cell.border = Border(
            left=make_side("left"), right=make_side("right"),
            top=make_side("top"), bottom=make_side("bottom"),
        )
    if "numberFormat" in style:
        cell.number_format = str(style["numberFormat"])
    if "protection" in style and isinstance(style["protection"], dict):
        from openpyxl.styles import Protection
        cell.protection = Protection(**style["protection"])


def _write_cell(cell, spec: object) -> None:
    if not isinstance(spec, dict):
        cell.value = spec
        return
    if "formula" in spec:
        formula = str(spec["formula"])
        cell.value = formula if formula.startswith("=") else f"={formula}"
    elif "value" in spec:
        cell.value = spec["value"]
    if isinstance(spec.get("style"), dict):
        _apply_style(cell, spec["style"])
    if "comment" in spec:
        from openpyxl.comments import Comment
        cell.comment = Comment(str(spec["comment"]), str(spec.get("commentAuthor", "Craft Agent")))
    if "hyperlink" in spec:
        cell.hyperlink = str(spec["hyperlink"])


def _write_matrix(ws, start: str, values: list[list[object]], style: dict[str, object] | None = None) -> None:
    start_row, start_column = coordinate_to_tuple(start.upper())
    for row_offset, row in enumerate(values):
        if not isinstance(row, list):
            raise ValueError("range values must be a two-dimensional JSON array")
        for column_offset, value in enumerate(row):
            cell = ws.cell(row=start_row + row_offset, column=start_column + column_offset)
            _write_cell(cell, value)
            if style:
                _apply_style(cell, style)


def _apply_sheet_spec(ws, spec: dict[str, object]) -> None:
    rows = spec.get("rows", [])
    if rows:
        if not isinstance(rows, list) or any(not isinstance(row, list) for row in rows):
            raise ValueError("sheet rows must be a two-dimensional JSON array")
        for row in rows:
            ws.append(row)

    cells = spec.get("cells", {})
    if not isinstance(cells, dict):
        raise ValueError("sheet cells must be an object keyed by A1 reference")
    for coordinate, cell_spec in cells.items():
        _write_cell(ws[str(coordinate).upper()], cell_spec)

    ranges = spec.get("ranges", [])
    if not isinstance(ranges, list):
        raise ValueError("sheet ranges must be an array")
    for range_spec in ranges:
        if not isinstance(range_spec, dict) or "start" not in range_spec or "values" not in range_spec:
            raise ValueError("each range requires start and values")
        style = range_spec.get("style")
        _write_matrix(
            ws,
            str(range_spec["start"]),
            range_spec["values"],
            style if isinstance(style, dict) else None,
        )

    for merged_range in spec.get("merges", []):
        ws.merge_cells(str(merged_range))
    if spec.get("freezePanes"):
        ws.freeze_panes = str(spec["freezePanes"])
    if spec.get("autoFilter"):
        ws.auto_filter.ref = str(spec["autoFilter"])

    column_widths = spec.get("columnWidths", {})
    if isinstance(column_widths, dict):
        for column, width in column_widths.items():
            ws.column_dimensions[str(column).upper()].width = float(width)
    row_heights = spec.get("rowHeights", {})
    if isinstance(row_heights, dict):
        for row, height in row_heights.items():
            ws.row_dimensions[int(row)].height = float(height)

    for rule_spec in spec.get("conditionalFormats", []):
        if not isinstance(rule_spec, dict) or "range" not in rule_spec:
            raise ValueError("conditional format requires range")
        rule_type = rule_spec.get("type", "colorScale")
        if rule_type == "colorScale":
            colors = rule_spec.get("colors", ["F8696B", "FFEB84", "63BE7B"])
            if not isinstance(colors, list) or len(colors) not in (2, 3):
                raise ValueError("colorScale colors must contain two or three colors")
            if len(colors) == 2:
                rule = ColorScaleRule(start_type="min", start_color=_color(colors[0]), end_type="max", end_color=_color(colors[1]))
            else:
                rule = ColorScaleRule(start_type="min", start_color=_color(colors[0]), mid_type="percentile", mid_value=50, mid_color=_color(colors[1]), end_type="max", end_color=_color(colors[2]))
        elif rule_type == "cellIs":
            rule = CellIsRule(
                operator=str(rule_spec.get("operator", "greaterThan")),
                formula=[str(value) for value in rule_spec.get("formula", ["0"])],
                fill=PatternFill(fill_type="solid", fgColor=_color(rule_spec.get("color"), "FFEB9C")),
            )
        elif rule_type == "formula":
            rule = FormulaRule(
                formula=[str(value) for value in rule_spec.get("formula", [])],
                fill=PatternFill(fill_type="solid", fgColor=_color(rule_spec.get("color"), "FFEB9C")),
            )
        else:
            raise ValueError(f"unsupported conditional format type: {rule_type}")
        ws.conditional_formatting.add(str(rule_spec["range"]), rule)

    for validation_spec in spec.get("dataValidations", []):
        if not isinstance(validation_spec, dict) or "range" not in validation_spec:
            raise ValueError("data validation requires range")
        validation = DataValidation(
            type=str(validation_spec.get("type", "list")),
            formula1=validation_spec.get("formula1"),
            formula2=validation_spec.get("formula2"),
            operator=validation_spec.get("operator"),
            allow_blank=bool(validation_spec.get("allowBlank", True)),
        )
        validation.error = validation_spec.get("error")
        validation.errorTitle = validation_spec.get("errorTitle")
        validation.prompt = validation_spec.get("prompt")
        validation.promptTitle = validation_spec.get("promptTitle")
        ws.add_data_validation(validation)
        validation.add(str(validation_spec["range"]))

    chart_classes = {"bar": BarChart, "column": BarChart, "line": LineChart, "pie": PieChart}
    for chart_spec in spec.get("charts", []):
        if not isinstance(chart_spec, dict):
            raise ValueError("chart must be an object")
        chart_type = str(chart_spec.get("type", "bar"))
        chart_class = chart_classes.get(chart_type)
        if chart_class is None:
            raise ValueError(f"unsupported chart type: {chart_type}")
        chart = chart_class()
        min_col, min_row, max_col, max_row = range_boundaries(str(chart_spec["data"]))
        chart.add_data(
            Reference(ws, min_col=min_col, min_row=min_row, max_col=max_col, max_row=max_row),
            titles_from_data=bool(chart_spec.get("titlesFromData", True)),
        )
        if chart_spec.get("categories"):
            c_min_col, c_min_row, c_max_col, c_max_row = range_boundaries(str(chart_spec["categories"]))
            chart.set_categories(Reference(ws, min_col=c_min_col, min_row=c_min_row, max_col=c_max_col, max_row=c_max_row))
        chart.title = chart_spec.get("title")
        chart.height = float(chart_spec.get("height", 7.5))
        chart.width = float(chart_spec.get("width", 15))
        if chart_type == "bar":
            chart.type = "bar"
        ws.add_chart(chart, str(chart_spec.get("anchor", "E2")))

    for image_spec in spec.get("images", []):
        if not isinstance(image_spec, dict) or "path" not in image_spec:
            raise ValueError("image requires path")
        image = SpreadsheetImage(str(image_spec["path"]))
        if image_spec.get("width") is not None:
            image.width = float(image_spec["width"])
        if image_spec.get("height") is not None:
            image.height = float(image_spec["height"])
        ws.add_image(image, str(image_spec.get("anchor", "A1")))


@click.group()
def cli() -> None:
    """Excel (.xlsx) operations tool."""
    pass


def _read_sheet_data(ws, cell_range: str | None = None) -> list[list[object]]:
    """Read data from a worksheet, returning list of rows."""
    if cell_range:
        rows = list(ws[cell_range])
    else:
        rows = list(ws.iter_rows())
    return [[cell.value for cell in row] for row in rows]


def _build_records(data: list[list[object]]) -> list[dict[str, object]]:
    """Convert row data (with header row) to list of dicts."""
    if not data or len(data) <= 1:
        return []
    headers = [str(h) if h is not None else f"col_{i}" for i, h in enumerate(data[0])]
    records = []
    for row_data in data[1:]:
        record: dict[str, object] = {}
        for i, val in enumerate(row_data):
            key = headers[i] if i < len(headers) else f"col_{i}"
            record[key] = val
        records.append(record)
    return records


def _format_data(data: list[list[object]], fmt: str) -> str:
    """Format row data as text, csv, or json."""
    if not data:
        if fmt == "json":
            return "[]"
        elif fmt == "csv":
            return ""
        else:
            return "(empty)"

    if fmt == "json":
        if len(data) > 1:
            headers = [str(h) if h is not None else f"col_{i}" for i, h in enumerate(data[0])]
            records = []
            for row_data in data[1:]:
                record: dict[str, object] = {}
                for i, val in enumerate(row_data):
                    key = headers[i] if i < len(headers) else f"col_{i}"
                    record[key] = val
                records.append(record)
            return json.dumps(records, indent=2, default=_json_serial)
        else:
            # Single row = header only, no data rows
            return "[]"
    elif fmt == "csv":
        buf = io.StringIO()
        writer = csv.writer(buf)
        for row_data in data:
            writer.writerow(row_data)
        return buf.getvalue()
    else:
        lines: list[str] = []
        str_data = [[str(v) if v is not None else "" for v in row] for row in data]
        if str_data:
            max_cols = max(len(row) for row in str_data)
            col_widths = [0] * max_cols
            for row in str_data:
                for i, val in enumerate(row):
                    col_widths[i] = max(col_widths[i], len(val))
            for row in str_data:
                parts = []
                for i, val in enumerate(row):
                    width = col_widths[i] if i < len(col_widths) else 0
                    parts.append(val.ljust(width))
                lines.append("  ".join(parts).rstrip())
        return "\n".join(lines)


@cli.command()
@click.argument("file", type=click.Path(exists=True, dir_okay=False))
@click.option("--sheet", type=str, default=None, help="Sheet name (default: active sheet).")
@click.option("--all-sheets", is_flag=True, default=False, help="Read all sheets in the workbook.")
@click.option("--range", "cell_range", type=str, default=None, help="Cell range, e.g. 'A1:C10'.")
@click.option("--format", "fmt", type=click.Choice(["text", "csv", "json"]), default="text", help="Output format.")
@click.option("-o", "--output", type=click.Path(), default=None, help="Write output to file.")
def read(file: str, sheet: str | None, all_sheets: bool, cell_range: str | None, fmt: str, output: str | None) -> None:
    """Read cells, ranges, or entire sheets from an Excel file."""
    try:
        wb = load_workbook(file, read_only=True, data_only=True)

        if all_sheets and sheet:
            wb.close()
            click.echo("Error: --all-sheets and --sheet are mutually exclusive.", err=True)
            sys.exit(1)

        if all_sheets:
            # Read all sheets
            if fmt == "json":
                all_data: dict[str, object] = {}
                for name in wb.sheetnames:
                    ws = wb[name]
                    data = _read_sheet_data(ws, cell_range)
                    all_data[name] = _build_records(data)
                result = json.dumps(all_data, indent=2, default=_json_serial)
            else:
                parts: list[str] = []
                for name in wb.sheetnames:
                    ws = wb[name]
                    data = _read_sheet_data(ws, cell_range)
                    if fmt == "csv":
                        parts.append(f"# Sheet: {name}")
                    else:
                        parts.append(f"=== Sheet: {name} ===")
                    parts.append(_format_data(data, fmt))
                    parts.append("")
                result = "\n".join(parts)
            wb.close()
            write_output(result, output)
            return

        if sheet:
            if sheet not in wb.sheetnames:
                wb.close()
                click.echo(f"Error: sheet '{sheet}' not found. Available: {', '.join(wb.sheetnames)}", err=True)
                sys.exit(1)
            ws = wb[sheet]
        else:
            ws = wb.active
            if ws is None:
                click.echo("Error: no active sheet found. Use --sheet to specify one.", err=True)
                wb.close()
                sys.exit(1)

        data = _read_sheet_data(ws, cell_range)
        result = _format_data(data, fmt)

        wb.close()
        write_output(result, output)
    except Exception as e:
        click.echo(f"Error: {e}", err=True)
        sys.exit(1)


@cli.command()
@click.argument("file", type=click.Path(dir_okay=False))
@click.option("--sheet", type=str, default=None, help="Sheet name (default: active sheet).")
@click.option("--cell", type=str, required=True, help="Cell reference, e.g. 'A1'.")
@click.option("--value", type=str, required=True, help="Value to write.")
@click.option("--type", "val_type", type=click.Choice(["string", "number", "bool"]), default="string", help="Value type.")
def write(file: str, sheet: str | None, cell: str, value: str, val_type: str) -> None:
    """Write a value to a specific cell in an Excel file.

    Creates the file if it does not exist.
    """
    try:
        file_path = Path(file)
        if file_path.exists():
            wb = load_workbook(file)
        else:
            wb = Workbook()

        if sheet:
            if sheet not in wb.sheetnames:
                wb.create_sheet(sheet)
            ws = wb[sheet]
        else:
            ws = wb.active
            if ws is None:
                click.echo("Error: no active sheet found. Use --sheet to specify one.", err=True)
                wb.close()
                sys.exit(1)

        # Convert value type
        converted: object
        if val_type == "number":
            try:
                converted = int(value)
            except ValueError:
                try:
                    converted = float(value)
                except ValueError:
                    click.echo(f"Error: '{value}' is not a valid number.", err=True)
                    wb.close()
                    sys.exit(1)
        elif val_type == "bool":
            converted = value.lower() in ("true", "1", "yes")
        else:
            converted = value

        ws[cell.upper()] = converted
        wb.save(file)
        wb.close()
        click.echo(f"Wrote '{converted}' to {cell.upper()} in {file_path.name}", err=True)
    except Exception as e:
        click.echo(f"Error: {e}", err=True)
        sys.exit(1)


@cli.command("write-range")
@click.argument("file", type=click.Path(dir_okay=False))
@click.option("--sheet", type=str, default=None, help="Sheet name (default: active sheet).")
@click.option("--start", type=str, required=True, help="Top-left cell, e.g. A1.")
@click.option("--values", type=str, required=True, help="Two-dimensional JSON array or path to a JSON file.")
@click.option("--style", type=str, default=None, help="Optional JSON cell style applied to the whole range.")
def write_range(file: str, sheet: str | None, start: str, values: str, style: str | None) -> None:
    """Write a batch matrix (including formulas and per-cell specs) in one operation."""
    try:
        parsed = _load_json_input(values)
        if not isinstance(parsed, list):
            raise ValueError("--values must be a two-dimensional JSON array")
        parsed_style = _load_json_input(style) if style else None
        if parsed_style is not None and not isinstance(parsed_style, dict):
            raise ValueError("--style must be a JSON object")
        path = Path(file)
        wb = load_workbook(path) if path.exists() else Workbook()
        if sheet:
            ws = wb[sheet] if sheet in wb.sheetnames else wb.create_sheet(sheet)
        else:
            ws = wb.active
            if ws is None:
                ws = wb.create_sheet("Sheet1")
        _write_matrix(ws, start, parsed, parsed_style)
        path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(path)
        wb.close()
        click.echo(f"Wrote {len(parsed)} row(s) starting at {start.upper()} in {path.name}", err=True)
    except (json.JSONDecodeError, ValueError) as e:
        click.echo(f"Error: {e}", err=True)
        sys.exit(1)
    except Exception as e:
        click.echo(f"Error: {e}", err=True)
        sys.exit(1)


@cli.command()
@click.option("--spec", type=str, required=True, help="Workbook JSON spec or path to a JSON spec file.")
@click.option("-o", "--output", type=click.Path(), required=True, help="Output .xlsx file path.")
def build(spec: str, output: str) -> None:
    """Build a styled workbook with ranges, formulas, validation, charts and images.

    The spec is an object with a ``sheets`` array. Each sheet supports ``rows``,
    ``cells``, ``ranges``, ``merges``, ``conditionalFormats``,
    ``dataValidations``, ``charts``, ``images``, dimensions and filters.
    """
    try:
        parsed = _load_json_input(spec)
        if not isinstance(parsed, dict) or not isinstance(parsed.get("sheets"), list) or not parsed["sheets"]:
            raise ValueError("workbook spec requires a non-empty sheets array")
        wb = Workbook()
        default_sheet = wb.active
        for index, sheet_spec in enumerate(parsed["sheets"]):
            if not isinstance(sheet_spec, dict):
                raise ValueError("each sheet spec must be an object")
            name = str(sheet_spec.get("name", f"Sheet{index + 1}"))
            if index == 0 and default_sheet is not None:
                ws = default_sheet
                ws.title = name
            else:
                if name in wb.sheetnames:
                    raise ValueError(f"duplicate sheet name: {name}")
                ws = wb.create_sheet(name)
            _apply_sheet_spec(ws, sheet_spec)

        properties = parsed.get("properties")
        if isinstance(properties, dict):
            for key in ("title", "subject", "creator", "description", "keywords", "category"):
                if key in properties:
                    setattr(wb.properties, key, str(properties[key]))
        output_path = Path(output)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(output_path)
        wb.close()
        click.echo(f"Workbook built: {output_path} ({len(parsed['sheets'])} sheets)", err=True)
    except (json.JSONDecodeError, ValueError, KeyError) as e:
        click.echo(f"Error: {e}", err=True)
        sys.exit(1)
    except Exception as e:
        click.echo(f"Error: {e}", err=True)
        sys.exit(1)


@cli.command()
@click.argument("file", type=click.Path(exists=True, dir_okay=False))
@click.option("-o", "--output", type=click.Path(), default=None, help="Write output to file.")
def info(file: str, output: str | None) -> None:
    """Show workbook information: sheets, dimensions, cell counts."""
    try:
        wb = load_workbook(file, read_only=False, data_only=False)

        sheets_info = []
        for name in wb.sheetnames:
            ws = wb[name]
            try:
                dimensions = ws.dimensions
            except Exception:
                # ReadOnlyWorksheet may not expose .dimensions in some openpyxl versions.
                dimensions = ws.calculate_dimension() if hasattr(ws, "calculate_dimension") else None

            formula_count = sum(
                1 for row in ws.iter_rows() for cell in row
                if isinstance(cell.value, str) and cell.value.startswith("=")
            )
            sheets_info.append({
                "name": name,
                "dimensions": dimensions,
                "min_row": ws.min_row,
                "max_row": ws.max_row,
                "min_column": ws.min_column,
                "max_column": ws.max_column,
                "formula_count": formula_count,
                "merged_ranges": [str(value) for value in ws.merged_cells.ranges],
                "conditional_format_count": len(ws.conditional_formatting),
                "data_validation_count": len(ws.data_validations.dataValidation),
                "chart_count": len(ws._charts),
                "image_count": len(ws._images),
                "freeze_panes": str(ws.freeze_panes) if ws.freeze_panes else None,
                "auto_filter": ws.auto_filter.ref,
            })

        info_dict = {
            "file": str(Path(file).resolve()),
            "sheet_count": len(wb.sheetnames),
            "sheet_names": wb.sheetnames,
            "sheets": sheets_info,
        }

        wb.close()
        write_output(json.dumps(info_dict, indent=2, default=_json_serial), output)
    except Exception as e:
        click.echo(f"Error: {e}", err=True)
        sys.exit(1)


@cli.command("add-sheet")
@click.argument("file", type=click.Path(dir_okay=False))
@click.option("--name", type=str, required=True, help="Name for the new sheet.")
@click.option("--position", type=int, default=None, help="Position index (0-based). Default: append at end.")
def add_sheet(file: str, name: str, position: int | None) -> None:
    """Add a new sheet to an Excel file.

    Creates the file if it does not exist.
    """
    try:
        file_path = Path(file)
        if file_path.exists():
            wb = load_workbook(file)
        else:
            wb = Workbook()
            # Remove default sheet if creating new file
            if "Sheet" in wb.sheetnames:
                del wb["Sheet"]

        if name in wb.sheetnames:
            wb.close()
            click.echo(f"Error: sheet '{name}' already exists.", err=True)
            sys.exit(1)

        if position is not None:
            wb.create_sheet(name, position)
        else:
            wb.create_sheet(name)

        wb.save(file)
        wb.close()
        click.echo(f"Added sheet '{name}' to {file_path.name}", err=True)
    except Exception as e:
        click.echo(f"Error: {e}", err=True)
        sys.exit(1)


@cli.command()
@click.argument("file", type=click.Path(exists=True, dir_okay=False))
@click.option("--sheet", type=str, default=None, help="Sheet name (default: active sheet).")
@click.option("--all-sheets", is_flag=True, default=False, help="Export all sheets in the workbook.")
@click.option("--format", "fmt", type=click.Choice(["csv", "json"]), default="csv", help="Export format.")
@click.option("-o", "--output", type=click.Path(), default=None, help="Write output to file.")
def export(file: str, sheet: str | None, all_sheets: bool, fmt: str, output: str | None) -> None:
    """Export a sheet as CSV or JSON."""
    try:
        wb = load_workbook(file, read_only=True, data_only=True)

        if all_sheets and sheet:
            wb.close()
            click.echo("Error: --all-sheets and --sheet are mutually exclusive.", err=True)
            sys.exit(1)

        if all_sheets:
            if fmt == "json":
                all_data: dict[str, object] = {}
                for name in wb.sheetnames:
                    ws = wb[name]
                    data = _read_sheet_data(ws)
                    all_data[name] = _build_records(data)
                result = json.dumps(all_data, indent=2, default=_json_serial)
            else:
                parts: list[str] = []
                for name in wb.sheetnames:
                    ws = wb[name]
                    data = _read_sheet_data(ws)
                    parts.append(f"# Sheet: {name}")
                    parts.append(_format_data(data, "csv"))
                    parts.append("")
                result = "\n".join(parts)
            wb.close()
            write_output(result, output)
            return

        if sheet:
            if sheet not in wb.sheetnames:
                wb.close()
                click.echo(f"Error: sheet '{sheet}' not found. Available: {', '.join(wb.sheetnames)}", err=True)
                sys.exit(1)
            ws = wb[sheet]
        else:
            ws = wb.active
            if ws is None:
                click.echo("Error: no active sheet found. Use --sheet to specify one.", err=True)
                wb.close()
                sys.exit(1)

        data = _read_sheet_data(ws)
        result = _format_data(data, fmt)

        wb.close()
        write_output(result, output)
    except Exception as e:
        click.echo(f"Error: {e}", err=True)
        sys.exit(1)


if __name__ == "__main__":
    cli()
